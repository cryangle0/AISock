"""
把 docs/客户准备清单.csv 转成带格式的 .xlsx：
  - 第 1 行表头加粗+棕色背景+白字
  - 列宽自适应
  - 单元格自动换行
  - 必需/可选列条件填色（必需=朱砂红淡底，可选=沙金淡底）
  - 冻结首行
  - 类别列合并相邻同类单元格
"""
from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "客户准备清单.csv"
OUT = ROOT / "客户准备清单.xlsx"

# Dunhuang theme colors
BROWN_DARK   = "8C5A3C"   # 标题棕
CREAM        = "FFFCF6"   # 表头字白
ROSE_SOFT    = "FBE5E0"   # 必需淡朱砂
SAND_SOFT    = "F5EBD0"   # 可选淡沙金
GRID         = "E5D8C0"   # 沙金边框

def main():
    rows = list(csv.reader(SRC.open(encoding="utf-8")))
    if not rows:
        raise RuntimeError("CSV is empty")

    wb = Workbook()
    ws = wb.active
    ws.title = "客户准备清单"

    # 头/列宽
    headers = rows[0]
    widths = [12, 24, 50, 24, 12, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    side = Side(style="thin", color=GRID)
    border = Border(left=side, right=side, top=side, bottom=side)

    # 写入数据
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left",
            )
            cell.border = border

    # 表头样式
    head_fill = PatternFill("solid", fgColor=BROWN_DARK)
    head_font = Font(name="微软雅黑", size=11, bold=True, color=CREAM)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal="center",
        )
    ws.row_dimensions[1].height = 28

    # 数据行：必需/可选 着色（第 5 列）
    rose_fill = PatternFill("solid", fgColor=ROSE_SOFT)
    sand_fill = PatternFill("solid", fgColor=SAND_SOFT)
    for r in range(2, len(rows) + 1):
        flag = ws.cell(row=r, column=5).value or ""
        flag = flag.strip()
        cell = ws.cell(row=r, column=5)
        if flag.startswith("必需"):
            cell.fill = rose_fill
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="C5483C")
        elif flag.startswith("推荐"):
            cell.fill = sand_fill
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="C69130")
        elif flag.startswith("可选"):
            cell.font = Font(name="微软雅黑", size=10, color="6B5A48")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[r].height = 36

    # 默认正文字体
    body_font = Font(name="微软雅黑", size=10, color="2B1F14")
    for r in range(2, len(rows) + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font.bold:
                continue
            cell.font = body_font

    # 合并第 1 列（类别）相邻同名单元格
    if len(rows) > 2:
        start_row = 2
        cur_val = ws.cell(row=2, column=1).value
        for r in range(3, len(rows) + 2):
            val = ws.cell(row=r, column=1).value if r <= len(rows) else None
            if val != cur_val:
                end_row = r - 1
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, start_column=1,
                                   end_row=end_row,   end_column=1)
                start_row = r
                cur_val = val

    # 类别列样式
    cat_font = Font(name="微软雅黑", size=11, bold=True, color=BROWN_DARK)
    for r in range(2, len(rows) + 1):
        c = ws.cell(row=r, column=1)
        c.font = cat_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.fill = PatternFill("solid", fgColor="F5EDE0")  # 米色暖底

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(OUT)
    print(f"OK -> {OUT}")

if __name__ == "__main__":
    main()
